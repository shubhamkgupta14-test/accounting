from datetime import date, datetime, timezone
from io import BytesIO
import json
import re
import zipfile

from bson import ObjectId
from fastapi import APIRouter, Depends, File, Form, HTTPException, Query, UploadFile
from openpyxl import Workbook, load_workbook
from pydantic import BaseModel, Field, ValidationError
from fastapi.responses import StreamingResponse
from typing import Literal
from pymongo import ReturnDocument

from app.core.database import get_database
from app.core.config import settings
from app.dependencies import get_current_user, require_roles
from app.schemas import AccountCreate, JournalEntryCreate
from app.utils import object_id, serialize_doc, serialize_many
from app.pagination import PageParams, SortOrder, page_response, safe_search, sort_direction
from app.financial_reports import build_financial_report, get_financial_year

router = APIRouter(prefix="/journal-entries", tags=["journal entries"])


async def _next_system_voucher_no(db) -> str:
    numbers = []
    for value in await db.journal_entries.distinct("voucher_no"):
        match = re.search(r"(\d+)$", str(value or ""))
        if match:
            numbers.append(int(match.group(1)))
    return f"SYS-{max(numbers, default=0) + 1:03d}"


async def _system_voucher(db, system_entry_type: str, period, legacy_voucher: str):
    existing = await db.journal_entries.find_one({
        "system_entry_type": system_entry_type,
        "$or": [
            {"financial_year_start": period.start_date.isoformat()},
            {"voucher_no": legacy_voucher},
        ],
    })
    voucher_no = (
        existing.get("voucher_no")
        if existing and str(existing.get("voucher_no", "")).startswith("SYS-")
        else await _next_system_voucher_no(db)
    )
    return existing, voucher_no


async def _post_profit_transfer(db, closing_date: date, current_user: dict, *, persist: bool = True, voucher_override: str | None = None, narration_override: str | None = None):
    """Create or refresh the FY profit-to-capital journal after stock closing."""
    period = get_financial_year(closing_date)
    statement = await build_financial_report(db, period)
    profit = round(float(statement["profit_and_loss"]["net_profit"]), 2)
    retirement_allocations = await db.journal_entries.aggregate([
        {"$match": {
            "date": {"$gte": period.start_date.isoformat(), "$lte": period.end_date.isoformat()},
            "status": "Posted", "system_entry_type": "RETIREMENT_PROFIT_TRANSFER",
        }},
        {"$unwind": "$entries"},
        {"$match": {"entries.account": "Profit & Loss Account"}},
        {"$group": {"_id": None, "debit": {"$sum": "$entries.debit"}, "credit": {"$sum": "$entries.credit"}}},
    ]).to_list(length=1)
    if retirement_allocations:
        allocated = float(retirement_allocations[0].get("debit", 0) or 0) - float(retirement_allocations[0].get("credit", 0) or 0)
        profit = round(profit - allocated, 2)
    legacy_voucher = f"PROFIT-TRANSFER-{period.start_date.year}-{str(period.end_date.year)[-2:]}"
    existing, voucher_no = await _system_voucher(db, "PROFIT_TRANSFER", period, legacy_voucher)
    if abs(profit) < .005:
        if persist and existing:
            await db.journal_entries.delete_one({"_id": existing["_id"]})
        return None

    saved_settings = await db.app_settings.find_one({"_id": "global"}, {"partners": 1}) or {}
    configured_partners = saved_settings.get("partners", [])
    capital_accounts = []
    for partner in configured_partners:
        admission_date = str(partner.get("admission_date") or "")
        retirement_date = str(partner.get("retirement_date") or "")
        share = float(partner.get("share_percentage", 0) or 0)
        if share <= 0 or (admission_date and admission_date > closing_date.isoformat()) or (
            retirement_date and retirement_date <= closing_date.isoformat()
        ):
            continue
        account = await db.accounts.find_one({
            "name": partner.get("account_name"), "type": "Equity", "group": "Capital"
        })
        if account:
            capital_accounts.append((account, share))
    if not capital_accounts:
        capital = await db.accounts.find_one({"name": "Capital", "type": "Equity"})
        if capital is None:
            capital = await db.accounts.find_one({
                "type": "Equity", "group": "Capital", "name": {"$ne": "Profit & Loss Account"}
            }, sort=[("code", 1)])
        if capital:
            capital_accounts = [(capital, 100.0)]
    if not capital_accounts:
        raise HTTPException(status_code=400, detail="Create a Capital equity account before closing stock")

    if persist:
        await db.accounts.update_one(
            {"name": "Profit & Loss Account"},
            {"$setOnInsert": {
                "code": "SYS-PNL", "name": "Profit & Loss Account", "type": "Equity",
                "group": "Capital", "opening_balance": 0.0, "is_active": True,
            }},
            upsert=True,
        )
    amount = abs(profit)
    allocations = []
    allocated = 0.0
    for index, (account, share) in enumerate(capital_accounts):
        share_amount = round(amount - allocated, 2) if index == len(capital_accounts) - 1 else round(amount * share / 100, 2)
        allocated += share_amount
        allocations.append((account["name"], share_amount))
    if profit > 0:
        entries = [{"account": "Profit & Loss Account", "debit": amount, "credit": 0.0}]
        entries.extend({"account": name, "debit": 0.0, "credit": value} for name, value in allocations)
    else:
        entries = [{"account": name, "debit": value, "credit": 0.0} for name, value in allocations]
        entries.append({"account": "Profit & Loss Account", "debit": 0.0, "credit": amount})
    now = datetime.now(timezone.utc)
    document = {
        "voucher_no": voucher_override or voucher_no,
        "date": closing_date.isoformat(),
        "narration": narration_override or f"Being net {'profit' if profit > 0 else 'loss'} for FY {period.start_date.year}-{str(period.end_date.year)[-2:]} transferred to Capital.",
        "entries": entries, "status": "Posted", "system_entry_type": "PROFIT_TRANSFER",
        "financial_year_start": period.start_date.isoformat(),
        "posted_by": current_user["id"], "posted_at": now,
    }
    if not persist:
        return document
    await db.journal_entries.update_one(
        ({"_id": existing["_id"]} if existing else {"voucher_no": voucher_no}),
        {"$set": document, "$setOnInsert": {"created_by": current_user["id"], "created_at": now}},
        upsert=True,
    )
    return document


async def _post_drawings_transfer(db, closing_date: date, current_user: dict, *, persist: bool = True, voucher_override: str | None = None, narration_override: str | None = None):
    """Close FY Drawings into the matching proprietor or partner Capital ledgers."""
    period = get_financial_year(closing_date)
    legacy_voucher = f"DRAWINGS-TRANSFER-{period.start_date.year}-{str(period.end_date.year)[-2:]}"
    existing, voucher_no = await _system_voucher(db, "DRAWINGS_TRANSFER", period, legacy_voucher)
    drawings_accounts = await db.accounts.find({
        "type": "Equity", "group": "Capital", "name": {"$regex": "drawings?", "$options": "i"}
    }).sort("code", 1).to_list(length=None)
    names = [account["name"] for account in drawings_accounts]
    if not names:
        if persist and existing:
            await db.journal_entries.delete_one({"_id": existing["_id"]})
        return None

    rows = await db.journal_entries.aggregate([
        {"$match": {
            "date": {"$gte": period.start_date.isoformat(), "$lte": period.end_date.isoformat()},
            "status": "Posted",
            "system_entry_type": {"$nin": ["FY_CLOSE", "DRAWINGS_TRANSFER"]},
        }},
        {"$unwind": "$entries"},
        {"$match": {"entries.account": {"$in": names}}},
        {"$group": {
            "_id": "$entries.account",
            "debit": {"$sum": "$entries.debit"},
            "credit": {"$sum": "$entries.credit"},
        }},
    ]).to_list(length=None)
    drawings = [
        (row["_id"], round(float(row.get("debit", 0) or 0) - float(row.get("credit", 0) or 0), 2))
        for row in rows
    ]
    drawings = [(name, amount) for name, amount in drawings if amount >= .005]
    total = round(sum(amount for _, amount in drawings), 2)
    if total < .005:
        if persist and existing:
            await db.journal_entries.delete_one({"_id": existing["_id"]})
        return None

    saved_settings = await db.app_settings.find_one({"_id": "global"}, {"partners": 1}) or {}
    configured_partners = saved_settings.get("partners", [])
    capital_totals: dict[str, float] = {}
    if configured_partners:
        def normalized(value: str) -> str:
            return "".join(character for character in value.casefold() if character.isalnum())

        for drawing_name, amount in drawings:
            drawing_key = normalized(drawing_name)
            matches = []
            for partner in configured_partners:
                partner_key = normalized(str(partner.get("partner_name", "")))
                capital_key = normalized(str(partner.get("account_name", "")).replace("Capital", ""))
                if (partner_key and partner_key in drawing_key) or (capital_key and capital_key in drawing_key):
                    matches.append(partner)
            # A single configured partner owns every drawings ledger in a sole-partner setup.
            if len(configured_partners) == 1 and not matches:
                matches = configured_partners
            if len(matches) != 1:
                raise HTTPException(
                    status_code=400,
                    detail=(
                        f"Unable to match Drawings account '{drawing_name}' to one partner. "
                        "Include the configured partner name in the Drawings account name."
                    ),
                )
            capital_name = str(matches[0].get("account_name", "")).strip()
            capital = await db.accounts.find_one({
                "name": capital_name, "type": "Equity", "group": "Capital",
            })
            if capital is None:
                raise HTTPException(
                    status_code=400,
                    detail=f"Create the configured partner Capital account '{capital_name}' before closing drawings",
                )
            capital_totals[capital_name] = round(capital_totals.get(capital_name, 0.0) + amount, 2)
    else:
        capital = await db.accounts.find_one({"name": "Capital", "type": "Equity"})
        if capital is None:
            raise HTTPException(status_code=400, detail="Create a Capital equity account before closing drawings")
        capital_totals[capital["name"]] = total

    entries = [
        {"account": capital_name, "debit": amount, "credit": 0.0}
        for capital_name, amount in capital_totals.items()
    ]
    entries.extend(
        {"account": name, "debit": 0.0, "credit": amount}
        for name, amount in drawings
    )
    now = datetime.now(timezone.utc)
    document = {
        "voucher_no": voucher_override or voucher_no,
        "date": period.end_date.isoformat(),
        "narration": narration_override or f"Being drawings for FY {period.start_date.year}-{str(period.end_date.year)[-2:]} transferred to the respective Capital account(s) at financial year end.",
        "entries": entries, "status": "Posted", "system_entry_type": "DRAWINGS_TRANSFER",
        "financial_year_start": period.start_date.isoformat(),
        "posted_by": current_user["id"], "posted_at": now,
    }
    if not persist:
        return document
    await db.journal_entries.update_one(
        ({"_id": existing["_id"]} if existing else {"voucher_no": voucher_no}),
        {"$set": document, "$setOnInsert": {"created_by": current_user["id"], "created_at": now}},
        upsert=True,
    )
    return document


def _is_closing_stock(payload: JournalEntryCreate) -> bool:
    closing_names = {"closing stock", "stock-in-hand", "stock in hand"}
    return any(line.account.strip().lower() in closing_names and line.debit > 0 for line in payload.entries)


class ClosingEntryOverride(BaseModel):
    system_entry_type: Literal["PROFIT_TRANSFER", "DRAWINGS_TRANSFER"]
    voucher_no: str = Field(min_length=1, max_length=50)
    narration: str = Field(min_length=1, max_length=500)


class ClosingConfirmation(BaseModel):
    closing_date: date
    entries: list[ClosingEntryOverride] = Field(max_length=2)


async def _require_closing_stock_entry(db, closing_date: date) -> None:
    exists = await db.journal_entries.find_one({
        "date": closing_date.isoformat(),
        "status": "Posted",
        "entries": {"$elemMatch": {
            "account": {"$regex": r"^(closing stock|stock-in-hand|stock in hand)$", "$options": "i"},
            "debit": {"$gt": 0},
        }},
    })
    if not exists:
        raise HTTPException(status_code=409, detail="Post the Closing Stock journal before confirming year-end entries")


def _excel_date(value) -> date:
    if isinstance(value, datetime):
        return value.date()
    if isinstance(value, date):
        return value
    text = str(value or "").strip()
    for pattern in ("%Y-%m-%d", "%d/%m/%Y", "%d-%m-%Y", "%m/%d/%Y"):
        try:
            return datetime.strptime(text, pattern).date()
        except ValueError:
            continue
    raise ValueError(f"Invalid date '{text}'")


def _header(value) -> str:
    return "_".join(str(value or "").strip().lower().replace("/", " ").replace("-", " ").split())


def _suggest_account(name: str, index: int) -> dict:
    text = name.lower()
    if any(word in text for word in ("sale", "income", "commission", "interest received")):
        account_type, group = "Income", "Indirect Income"
    elif any(word in text for word in ("expense", "purchase", "rent", "salary", "wages", "charges")):
        account_type, group = "Expense", "Indirect Expenses"
    elif any(word in text for word in ("payable", "creditor", "loan", "liability")):
        account_type, group = "Liability", "Current Liabilities"
    elif any(word in text for word in ("capital", "drawings", "equity")):
        account_type, group = "Equity", "Capital"
    else:
        account_type, group = "Asset", "Current Assets"
    return {"source_name": name, "name": name, "code": f"IMP-{index:03d}", "type": account_type, "group": group}


def _excel_account_names(content: bytes) -> set[str]:
    workbook = load_workbook(BytesIO(content), read_only=True, data_only=True)
    sheet = workbook.active
    _validate_sheet_dimensions(sheet)
    rows = sheet.iter_rows(values_only=True)
    headers = [_header(value) for value in next(rows)]
    candidates = {"account", "account_name", "ledger", "ledger_name"}
    column = next((headers.index(name) for name in candidates if name in headers), None)
    if column is None:
        raise ValueError("Missing Excel column: account")
    names = set()
    for row_number, row in enumerate(rows, start=2):
        if row_number > settings.max_excel_rows + 1:
            raise ValueError(
                f"Workbook exceeds the {settings.max_excel_rows} row limit")
        if column < len(row) and row[column] not in (None, ""):
            names.add(str(row[column]).strip())
    return names


def _validate_excel_container(content: bytes) -> None:
    if not zipfile.is_zipfile(BytesIO(content)):
        raise HTTPException(status_code=400, detail="The uploaded file is not a valid Excel workbook")
    try:
        with zipfile.ZipFile(BytesIO(content)) as archive:
            total_size = sum(item.file_size for item in archive.infolist())
            if total_size > settings.max_excel_uncompressed_bytes:
                raise HTTPException(
                    status_code=413,
                    detail="The uncompressed workbook is too large",
                )
            if len(archive.infolist()) > 2_000:
                raise HTTPException(
                    status_code=413,
                    detail="The workbook contains too many internal files",
                )
    except zipfile.BadZipFile as exc:
        raise HTTPException(status_code=400, detail="The uploaded file is not a valid Excel workbook") from exc


async def _read_excel_upload(file: UploadFile) -> bytes:
    content = await file.read(settings.max_excel_upload_bytes + 1)
    if len(content) > settings.max_excel_upload_bytes:
        raise HTTPException(
            status_code=413,
            detail=f"Excel uploads are limited to {settings.max_excel_upload_bytes} bytes",
        )
    _validate_excel_container(content)
    return content


def _validate_sheet_dimensions(sheet) -> None:
    if sheet.max_column > settings.max_excel_columns:
        raise HTTPException(
            status_code=422,
            detail=f"Workbook exceeds the {settings.max_excel_columns} column limit",
        )
    if sheet.max_row > settings.max_excel_rows + 1:
        raise HTTPException(
            status_code=422,
            detail=f"Workbook exceeds the {settings.max_excel_rows} data row limit",
        )


async def _validate_accounts(payload: JournalEntryCreate) -> None:
    db = get_database()
    account_names = {line.account for line in payload.entries}
    existing_accounts = await db.accounts.find(
        {"name": {"$in": list(account_names)}}
    ).to_list(500)
    existing_names = {account["name"] for account in existing_accounts}
    missing_accounts = sorted(account_names - existing_names)
    if missing_accounts:
        raise HTTPException(
            status_code=400,
            detail=f"Create/select valid accounts before posting this journal entry: {', '.join(missing_accounts)}",
        )


@router.get("")
async def list_journal_entries(_: dict = Depends(get_current_user)):
    docs = await get_database().journal_entries.find({}).sort("date", -1).to_list(500)
    return serialize_many(docs)


@router.get("/page")
async def page_journal_entries(
    params: PageParams = Depends(),
    search: str | None = Query(default=None, max_length=200),
    status_filter: Literal["Draft", "Posted"] | None = Query(default=None, alias="status"),
    date_from: str | None = None,
    date_to: str | None = None,
    sort_by: Literal["date", "voucher_no"] = "date",
    sort_order: SortOrder = "desc",
    _: dict = Depends(get_current_user),
):
    query: dict = {}
    pattern = safe_search(search)
    if pattern:
        query["$or"] = [{"voucher_no": {"$regex": pattern, "$options": "i"}}, {"narration": {"$regex": pattern, "$options": "i"}}]
    if status_filter:
        query["status"] = status_filter
    if date_from or date_to:
        query["date"] = {**({"$gte": date_from} if date_from else {}), **({"$lte": date_to} if date_to else {})}
    db = get_database()
    total = await db.journal_entries.count_documents(query)
    direction = sort_direction(sort_order)
    sort_fields = [(sort_by, direction)]
    # Accounting date controls chronology. _id is only a same-date tie-breaker
    # so the most recently inserted entry for that date appears first.
    sort_fields.append(("_id", -1))
    docs = await db.journal_entries.find(query).sort(sort_fields).skip(params.skip).limit(params.page_size).to_list(params.page_size)
    return page_response(docs, params, total)


@router.post("", status_code=201)
async def create_journal_entry(payload: JournalEntryCreate, current_user=Depends(require_roles("superadmin", "admin"))):
    db = get_database()
    if await db.journal_entries.find_one({"voucher_no": payload.voucher_no}):
        raise HTTPException(status_code=409, detail="Voucher number already exists")
    await _validate_accounts(payload)
    doc = payload.model_dump(mode="json")
    doc["status"] = "Posted"
    doc["created_by"] = current_user["id"]
    doc["created_at"] = datetime.now(timezone.utc)
    doc["posted_by"] = current_user["id"]
    doc["posted_at"] = datetime.now(timezone.utc)
    result = await db.journal_entries.insert_one(doc)
    return serialize_doc(await db.journal_entries.find_one({"_id": result.inserted_id}))


@router.post("/import-excel", status_code=201)
async def import_journal_entries_excel(
    file: UploadFile = File(...),
    account_definitions: str | None = Form(default=None),
    current_user=Depends(require_roles("superadmin", "admin")),
):
    if not file.filename or not file.filename.lower().endswith((".xlsx", ".xlsm")):
        raise HTTPException(status_code=400, detail="Upload an Excel .xlsx or .xlsm file")
    content = await _read_excel_upload(file)
    try:
        raw_definitions = json.loads(account_definitions) if account_definitions else []
        if not isinstance(raw_definitions, list):
            raise ValueError("account_definitions must be a list")
        if len(raw_definitions) > settings.max_excel_new_accounts:
            raise ValueError(
                f"At most {settings.max_excel_new_accounts} new ledgers may be created per import")
        definitions = []
        for raw in raw_definitions:
            if not isinstance(raw, dict):
                raise ValueError("Every account definition must be an object")
            source_name = str(raw.get("source_name", "")).strip()
            if not source_name or len(source_name) > 200:
                raise ValueError("Every new ledger requires a valid source name")
            account = AccountCreate(
                code=raw.get("code"),
                name=raw.get("name"),
                type=raw.get("type"),
                group=raw.get("group"),
                opening_balance=0,
                is_active=True,
            )
            definitions.append({
                "source_name": source_name,
                **account.model_dump(),
            })
        workbook = load_workbook(BytesIO(content), read_only=True, data_only=True)
        sheet = workbook.active
        _validate_sheet_dimensions(sheet)
        rows = sheet.iter_rows(values_only=True)
        headers = [_header(value) for value in next(rows)]
    except ValidationError as exc:
        raise HTTPException(status_code=422, detail=[
            error["msg"] for error in exc.errors()
        ][:20]) from exc
    except (json.JSONDecodeError, StopIteration, ValueError, OSError) as exc:
        raise HTTPException(status_code=400, detail="Unable to read the Excel workbook") from exc
    mappings = {str(item.get("source_name", "")).strip(): item for item in definitions if isinstance(item, dict)}

    aliases = {
        "voucher_no": {"voucher_no", "voucher", "voucher_number", "voucher_no."},
        "date": {"date", "entry_date", "voucher_date"},
        "narration": {"narration", "description", "particulars"},
        "account": {"account", "account_name", "ledger", "ledger_name"},
        "debit": {"debit", "dr", "debit_amount"},
        "credit": {"credit", "cr", "credit_amount"},
    }
    columns = {field: next((headers.index(name) for name in names if name in headers), None) for field, names in aliases.items()}
    missing = [field for field, index in columns.items() if index is None]
    if missing:
        raise HTTPException(status_code=400, detail=f"Missing Excel columns: {', '.join(missing)}")

    grouped: dict[str, dict] = {}
    current_voucher = ""
    errors: list[str] = []
    for row_number, values in enumerate(rows, start=2):
        if row_number > settings.max_excel_rows + 1:
            errors.append(
                f"Workbook exceeds the {settings.max_excel_rows} data row limit")
            break
        if not any(value not in (None, "") for value in values):
            continue
        def cell(field: str):
            index = columns[field]
            return values[index] if index is not None and index < len(values) else None
        voucher = str(cell("voucher_no") or "").strip() or current_voucher
        if not voucher:
            errors.append(f"Row {row_number}: Voucher No. is required")
            continue
        current_voucher = voucher
        item = grouped.setdefault(voucher, {"date": None, "narration": "", "entries": [], "rows": []})
        try:
            if cell("date") not in (None, ""):
                parsed_date = _excel_date(cell("date"))
                if item["date"] and item["date"] != parsed_date:
                    raise ValueError("different dates used for the same voucher")
                item["date"] = parsed_date
            narration = str(cell("narration") or "").strip()
            if narration:
                item["narration"] = narration
            source_account = str(cell("account") or "").strip()
            account = str(mappings.get(source_account, {}).get("name", source_account)).strip()
            debit = float(cell("debit") or 0)
            credit = float(cell("credit") or 0)
            item["entries"].append({"account": account, "debit": debit, "credit": credit})
            item["rows"].append(row_number)
        except (TypeError, ValueError) as exc:
            errors.append(f"Row {row_number}: {exc}")

    payloads: list[JournalEntryCreate] = []
    for voucher, item in grouped.items():
        if not item["date"]:
            errors.append(f"Voucher {voucher}: Date is required on its first row")
            continue
        if not item["narration"]:
            errors.append(f"Voucher {voucher}: Narration is required")
            continue
        try:
            payloads.append(JournalEntryCreate(
                date=item["date"], voucher_no=voucher, narration=item["narration"],
                entries=item["entries"], status="Posted",
            ))
        except ValidationError as exc:
            message = "; ".join(error["msg"] for error in exc.errors())
            errors.append(f"Voucher {voucher}: {message}")
            continue
        debit = sum(line.debit for line in payloads[-1].entries)
        credit = sum(line.credit for line in payloads[-1].entries)
        if abs(debit - credit) >= .005 or debit <= 0:
            payloads.pop()
            errors.append(f"Voucher {voucher}: Debit {debit:g} and Credit {credit:g} must match")

    db = get_database()
    voucher_numbers = [payload.voucher_no for payload in payloads]
    duplicates_in_file = sorted({number for number in voucher_numbers if voucher_numbers.count(number) > 1})
    existing = await db.journal_entries.distinct("voucher_no", {"voucher_no": {"$in": voucher_numbers}})
    account_names = {line.account for payload in payloads for line in payload.entries}
    saved_accounts = set(await db.accounts.distinct("name", {"name": {"$in": list(account_names)}}))
    proposed_accounts = {str(item.get("name", "")).strip() for item in definitions}
    missing_accounts = sorted(account_names - saved_accounts - proposed_accounts)
    definition_codes = [item["code"] for item in definitions]
    definition_names = [item["name"] for item in definitions]
    definition_sources = [item["source_name"] for item in definitions]
    existing_codes = set(await db.accounts.distinct("code", {"code": {"$in": definition_codes}}))
    existing_names = set(await db.accounts.distinct("name", {"name": {"$in": definition_names}}))
    if duplicates_in_file:
        errors.append(f"Duplicate vouchers in file: {', '.join(duplicates_in_file)}")
    if existing:
        errors.append(f"Voucher numbers already exist: {', '.join(sorted(existing))}")
    if missing_accounts:
        errors.append(f"Unknown accounts: {', '.join(missing_accounts)}")
    if len(definition_codes) != len(set(code.casefold() for code in definition_codes)):
        errors.append("New ledger codes must be unique within the import")
    if len(definition_names) != len(set(name.casefold() for name in definition_names)):
        errors.append("New ledger names must be unique within the import")
    if len(definition_sources) != len(set(name.casefold() for name in definition_sources)):
        errors.append("Excel source ledger names must be unique within the import")
    if existing_codes:
        errors.append(f"Ledger codes already exist: {', '.join(sorted(existing_codes))}")
    if existing_names:
        errors.append(f"Ledger names already exist: {', '.join(sorted(existing_names))}")
    if errors:
        raise HTTPException(status_code=422, detail=errors[:50])
    if not payloads:
        raise HTTPException(status_code=400, detail="The workbook contains no journal entries")

    now = datetime.now(timezone.utc)
    if definitions:
        await db.accounts.insert_many([{
            "code": str(item["code"]).strip(), "name": str(item["name"]).strip(),
            "type": item["type"], "group": str(item["group"]).strip(),
            "opening_balance": 0.0, "is_active": True, "created_at": now,
        } for item in definitions])
    documents = []
    for payload in payloads:
        document = payload.model_dump(mode="json")
        document.update({"status": "Posted", "created_by": current_user["id"], "created_at": now, "posted_by": current_user["id"], "posted_at": now, "import_source": file.filename})
        documents.append(document)
    await db.journal_entries.insert_many(documents)
    return {"imported": len(documents), "voucher_numbers": voucher_numbers, "line_count": sum(len(item.entries) for item in payloads)}


@router.post("/import-excel/preview")
async def preview_journal_excel(
    file: UploadFile = File(...),
    _: dict = Depends(require_roles("superadmin", "admin")),
):
    if not file.filename or not file.filename.lower().endswith((".xlsx", ".xlsm")):
        raise HTTPException(status_code=400, detail="Upload an Excel .xlsx or .xlsm file")
    try:
        names = _excel_account_names(await _read_excel_upload(file))
    except (StopIteration, ValueError, OSError) as exc:
        raise HTTPException(status_code=400, detail=str(exc) or "Unable to read workbook") from exc
    existing = set(await get_database().accounts.distinct("name", {"name": {"$in": list(names)}}))
    unknown = sorted(names - existing)
    return {"unknown_ledgers": [_suggest_account(name, index) for index, name in enumerate(unknown, start=1)]}


@router.get("/closing-preview")
async def preview_closing_entries(
    closing_date: date,
    current_user=Depends(require_roles("superadmin", "admin")),
):
    db = get_database()
    await _require_closing_stock_entry(db, closing_date)
    candidates = [
        await _post_profit_transfer(db, closing_date, current_user, persist=False),
        await _post_drawings_transfer(db, closing_date, current_user, persist=False),
    ]
    candidates = [item for item in candidates if item]
    next_number = int((await _next_system_voucher_no(db)).split("-")[-1])
    for index, item in enumerate(candidates):
        existing = await db.journal_entries.find_one({
            "system_entry_type": item["system_entry_type"],
            "financial_year_start": item["financial_year_start"],
        })
        item["voucher_no"] = existing.get("voucher_no") if existing else f"SYS-{next_number + index:03d}"
        item.pop("posted_by", None)
        item.pop("posted_at", None)
    return {"entries": candidates}


@router.get("/pending-closing-preview")
async def pending_closing_entries(
    current_user=Depends(require_roles("superadmin", "admin")),
):
    """Recover the latest year-end preview when Closing Stock was posted but transfers were not confirmed."""
    db = get_database()
    stock_names = re.compile(r"^(closing stock|stock-in-hand|stock in hand)$", re.IGNORECASE)
    closing_journals = await db.journal_entries.find({
        "status": "Posted",
        "entries": {"$elemMatch": {"account": stock_names, "debit": {"$gt": 0}}},
    }, {"date": 1}).sort("date", -1).to_list(length=50)
    for closing_date_value in dict.fromkeys(str(row.get("date", ""))[:10] for row in closing_journals):
        try:
            closing_date = date.fromisoformat(closing_date_value)
        except ValueError:
            continue
        period = get_financial_year(closing_date)
        confirmed_types = set(await db.journal_entries.distinct("system_entry_type", {
            "financial_year_start": period.start_date.isoformat(),
            "system_entry_type": {"$in": ["PROFIT_TRANSFER", "DRAWINGS_TRANSFER"]},
        }))
        candidates = []
        if "PROFIT_TRANSFER" not in confirmed_types:
            candidate = await _post_profit_transfer(db, closing_date, current_user, persist=False)
            if candidate:
                candidates.append(candidate)
        if "DRAWINGS_TRANSFER" not in confirmed_types:
            candidate = await _post_drawings_transfer(db, closing_date, current_user, persist=False)
            if candidate:
                candidates.append(candidate)
        if not candidates:
            continue
        next_number = int((await _next_system_voucher_no(db)).split("-")[-1])
        for index, item in enumerate(candidates):
            item["voucher_no"] = f"SYS-{next_number + index:03d}"
            item.pop("posted_by", None)
            item.pop("posted_at", None)
        return {"closing_date": closing_date_value, "entries": candidates}
    return {"closing_date": None, "entries": []}


@router.post("/closing-confirm")
async def confirm_closing_entries(
    payload: ClosingConfirmation,
    current_user=Depends(require_roles("superadmin", "admin")),
):
    db = get_database()
    await _require_closing_stock_entry(db, payload.closing_date)
    requested = {item.system_entry_type: item for item in payload.entries}
    if len(requested) != len(payload.entries):
        raise HTTPException(status_code=422, detail="Each closing entry type can be confirmed only once")
    financial_year_start = get_financial_year(payload.closing_date).start_date.isoformat()
    for item in payload.entries:
        duplicate = await db.journal_entries.find_one({"voucher_no": item.voucher_no})
        belongs_to_same_entry = duplicate and (
            duplicate.get("system_entry_type") == item.system_entry_type
            and duplicate.get("financial_year_start") == financial_year_start
        )
        if duplicate and not belongs_to_same_entry:
            raise HTTPException(status_code=409, detail=f"Voucher number {item.voucher_no} already exists")
    saved = []
    if "PROFIT_TRANSFER" in requested:
        item = requested["PROFIT_TRANSFER"]
        result = await _post_profit_transfer(
            db, payload.closing_date, current_user,
            voucher_override=item.voucher_no, narration_override=item.narration,
        )
        if result:
            saved.append(result)
    if "DRAWINGS_TRANSFER" in requested:
        item = requested["DRAWINGS_TRANSFER"]
        result = await _post_drawings_transfer(
            db, payload.closing_date, current_user,
            voucher_override=item.voucher_no, narration_override=item.narration,
        )
        if result:
            saved.append(result)
    return {"created": len(saved), "entries": saved}


@router.get("/import-excel/sample")
async def journal_import_sample(_: dict = Depends(get_current_user)):
    workbook = Workbook()
    sheet = workbook.active
    sheet.title = "Journal Entries"
    sheet.append(["Voucher No", "Date", "Narration", "Account", "Debit", "Credit"])
    sheet.append(["SAMPLE-2-LINE", "2026-04-01", "Two-line journal entry sample", "Cash", 1000, 0])
    sheet.append(["", "", "", "Sales", 0, 1000])
    sheet.append(["SAMPLE-3-LINE", "2026-04-02", "Three-line journal entry sample", "Cash", 600, 0])
    sheet.append(["", "", "", "Bank Account", 400, 0])
    sheet.append(["", "", "", "Capital", 0, 1000])
    sheet.freeze_panes = "A2"
    widths = {"A": 20, "B": 14, "C": 34, "D": 24, "E": 14, "F": 14}
    for column, width in widths.items():
        sheet.column_dimensions[column].width = width
    content = BytesIO()
    workbook.save(content)
    content.seek(0)
    return StreamingResponse(
        content,
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={"Content-Disposition": 'attachment; filename="journal-entry-import-sample.xlsx"'},
    )


@router.put("/{entry_id}")
async def update_journal_entry(
    entry_id: str,
    payload: JournalEntryCreate,
    current_user=Depends(require_roles("superadmin", "admin")),
):
    db = get_database()
    entry_object_id = object_id(entry_id, "Journal entry")
    existing = await db.journal_entries.find_one({"_id": entry_object_id})
    if not existing:
        raise HTTPException(status_code=404, detail="Journal entry not found")
    duplicate = await db.journal_entries.find_one({
        "voucher_no": payload.voucher_no,
        "_id": {"$ne": entry_object_id},
    })
    if duplicate:
        raise HTTPException(status_code=409, detail="Voucher number already exists")
    await _validate_accounts(payload)

    values = payload.model_dump(mode="json")
    values["status"] = "Posted"
    values.update({
        "updated_by": current_user["id"],
        "updated_at": datetime.now(timezone.utc),
    })
    update: dict = {"$set": values}
    if existing.get("status") != "Posted":
        values.update({
            "posted_by": current_user["id"],
            "posted_at": datetime.now(timezone.utc),
        })
    result = await db.journal_entries.find_one_and_update(
        {"_id": entry_object_id}, update, return_document=ReturnDocument.AFTER
    )
    return serialize_doc(result)


@router.delete("/{entry_id}", status_code=204)
async def delete_journal_entry(
    entry_id: str,
    _: dict = Depends(require_roles("superadmin", "admin")),
):
    db = get_database()
    entry_object_id = object_id(entry_id, "Journal entry")
    result = await db.journal_entries.delete_one({"_id": entry_object_id})
    if result.deleted_count == 0:
        raise HTTPException(status_code=404, detail="Journal entry not found")


@router.patch("/{entry_id}/post")
async def post_journal_entry(entry_id: str, _: dict = Depends(require_roles("superadmin", "admin"))):
    result = await get_database().journal_entries.find_one_and_update(
        {"_id": object_id(entry_id, "Journal entry"), "status": "Draft"}, {"$set": {"status": "Posted", "posted_at": datetime.now(timezone.utc)}}, return_document=ReturnDocument.AFTER
    )
    if not result:
        raise HTTPException(status_code=404, detail="Journal entry not found")
    return serialize_doc(result)
