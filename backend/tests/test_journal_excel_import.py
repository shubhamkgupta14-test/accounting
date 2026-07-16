from io import BytesIO
import json

from openpyxl import Workbook, load_workbook


def test_excel_import_supports_four_line_journal(client, login):
    login("admin")
    workbook = Workbook()
    sheet = workbook.active
    sheet.append(["Voucher No", "Date", "Narration", "Account", "Debit", "Credit"])
    sheet.append(["IMPORT-4-LINES", "2026-07-14", "Four-line Excel import", "Cash", 60, 0])
    sheet.append(["", "", "", "Bank Account", 40, 0])
    sheet.append(["", "", "", "Sales", 0, 90])
    sheet.append(["", "", "", "Capital", 0, 10])
    content = BytesIO()
    workbook.save(content)

    response = client.post(
        "/api/journal-entries/import-excel",
        files={"file": ("journals.xlsx", content.getvalue(), "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet")},
    )

    assert response.status_code == 201, response.text
    assert response.json()["imported"] == 1
    assert response.json()["line_count"] == 4
    imported = next(row for row in client.get("/api/journal-entries").json() if row["voucher_no"] == "IMPORT-4-LINES")
    assert len(imported["entries"]) == 4


def test_journal_import_sample_contains_two_and_three_line_entries(client, login):
    login()
    response = client.get("/api/journal-entries/import-excel/sample")
    assert response.status_code == 200
    sheet = load_workbook(BytesIO(response.content), read_only=True, data_only=True).active
    rows = list(sheet.iter_rows(values_only=True))
    assert rows[0] == ("Voucher No", "Date", "Narration", "Account", "Debit", "Credit")
    assert sum(1 for row in rows[1:] if row[0] == "SAMPLE-2-LINE" or (row[0] is None and row[3] == "Sales")) == 2
    assert sum(1 for row in rows[1:] if row[0] == "SAMPLE-3-LINE" or (row[0] is None and row[3] in {"Bank Account", "Capital"})) == 3


def test_excel_preview_and_import_create_unknown_ledger(client, login):
    login("admin")
    workbook = Workbook()
    sheet = workbook.active
    sheet.append(["Voucher No", "Date", "Narration", "Account", "Debit", "Credit"])
    sheet.append(["IMPORT-NEW-LEDGER", "2026-07-15", "Unknown ledger import", "Excel Testing Expense", 75, 0])
    sheet.append(["", "", "", "Cash", 0, 75])
    content = BytesIO()
    workbook.save(content)
    upload = ("unknown-ledger.xlsx", content.getvalue(), "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet")

    preview = client.post("/api/journal-entries/import-excel/preview", files={"file": upload})
    assert preview.status_code == 200
    assert preview.json()["unknown_ledgers"][0]["source_name"] == "Excel Testing Expense"

    definitions = [{"source_name": "Excel Testing Expense", "name": "Imported Testing Expense", "code": "IMP-TEST-EXP", "type": "Expense", "group": "Indirect Expenses"}]
    response = client.post(
        "/api/journal-entries/import-excel",
        files={"file": upload},
        data={"account_definitions": json.dumps(definitions)},
    )
    assert response.status_code == 201, response.text

    async def find_account():
        from app.core.database import get_database
        return await get_database().accounts.find_one({"name": "Imported Testing Expense"})
    account = client.portal.call(find_account)
    assert account["code"] == "IMP-TEST-EXP"
